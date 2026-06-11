# -*- coding: utf-8 -*-
"""Apply prices from technomer.xlsx to prisma/seed.ts"""
from __future__ import annotations

import json
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = os.path.join(os.path.dirname(__file__), "..")
SEED_PATH = os.path.join(ROOT, "prisma", "seed.ts")
XLSX_PATH = os.path.join(ROOT, "public", "прайсы", "техномер.xlsx")

CABLE_PREFIXES = [
    ("Кабель БПЭК-ЕК", "bpek-ek-cable"),
    ("Кабель БПЭК-СМТ/ТМ-07", "bpek-smt-tm07-cable"),
    ("Кабель БПЭК-ТС", "bpek-ts-cable"),
    ("Кабель БПЭК-Флоугаз", "bpek-flowgaz-cable"),
    ("Кабель БПЭК- Флоугаз", "bpek-flowgaz-cable"),
    ("Кабель  БПЭК-ВКГ", "bpek-vkg-cable"),
    ("Кабель  БПЭК-СПГ/Ирвис", "bpek-spg-irvis-cable"),
]

SLUG_BY_EXCEL: dict[str, str] = {
    "Счетчик газа СМТ-Смарт G4": "smt-smart",
    "Счетчик газа СМТ-Смарт G6": "smt-smart",
    "Счетчик газа СМТ-Смарт G10": "smt-smart",
    "Счетчики газа микротермальный СМТ-Смарт-К G4": "smt-smart-k",
    "Счетчики газа микротермальный  СМТ-Смарт-К G6": "smt-smart-k",
    "Счетчики газа микротермальный СМТ-Смарт-ДКЗ G4": "smt-smart-dkz",
    "Счетчики газа микротермальный  СМТ-Смарт-ДКЗ G6": "smt-smart-dkz",
    "Счетчик газа микротермальный СМТ-Комплекс G4": "smt-kompleks",
    "Счетчик газа микротермальный СМТ-Комплекс G6": "smt-kompleks",
    "Счетчик газа микротермальный СМТ-Комплекс G10": "smt-kompleks",
    "Счетчик газа микротермальный СМТ-Комплекс G16": "smt-kompleks",
    "Счетчик газа микротермальный СМТ-Комплекс G25": "smt-kompleks",
    "Счетчик газа микротермальный СМТ-Комплекс G40": "smt-kompleks-g40",
    "Счетчик газа микротермальный СМТ-Комплекс G40-2": "smt-kompleks-g40",
    "Счетчик газа микротермальный СМТ-Комплекс G65": "smt-kompleks-g65-g100",
    "Счетчик газа микротермальный СМТ-Комплекс G100": "smt-kompleks-g65-g100",
    "Счетчик газа микротермальный СМТ-Комплекс-К G4": "smt-kompleks-k",
    "Счетчик газа микротермальный СМТ-Комплекс-К G6": "smt-kompleks-k",
    "Счетчик газа микротермальный СМТ-Комплекс-К G10": "smt-kompleks-k",
    "Счетчик газа микротермальный СМТ-Комплекс-К G16": "smt-kompleks-k",
    "Счетчик газа микротермальный СМТ-Комплекс-К G25": "smt-kompleks-k",
    "Цифровой коммуникационный блок БПЭК-02/ЦК": "bpek-02-ck",
    "Цифровой коммуникационный блок БПЭК-02/ЦК-Ультра": "bpek-02-ck-ultra",
    "Цифровой коммуникационный блок БПЭК-03/ЦК": "bpek-03-ck",
    "Цифровой коммуникационный блок БПЭК-05/ЦК": "bpek-05-ck",
    "Цифровой коммуникационный блок БПЭК-04/ЦК-Ех": "bpek-04-ck-ex",
    "ПО «Газсеть: Стандарт»": "gazset-standart",
    "Датчик импульсов IN-Z61 0,65м": "in-z61-065m",
    "Датчик импульсов IN-Z61 2.5м": "in-z61-25m",
    "Датчик импульсов IN-S10 0,8м": "in-s10-08m",
    "Датчик импульсов IN-S10 2.5м": "in-s10-25m",
    "Кабель для выносного монтажа GSM-антенны 8м": "gsm-antenna-cable-8m",
    "Антенна выносная GSM 3 м.": "gsm-external-antenna-3m",
    "Устройство считывающее оптическое КАО-USB": "kao-usb",
    "Элементы питания": "elementy-pitaniya",
    "Шкаф защитный ШГ-1": "shkaf-zashchitnyy-shg",
    "Шкаф защитный ШГ-2": "shkaf-zashchitnyy-shg",
    "Шкаф защитный ШГ-3": "shkaf-zashchitnyy-shg",
}


def norm_name(name: str) -> str:
    return re.sub(r"\s+", " ", name.strip())


def parse_length_m(name: str) -> int | None:
    m = re.search(r"\((\d+)\s*[Мм]\.?\)", name)
    if m:
        return int(m.group(1))
    m = re.search(r"\((\d+)м\)", name, re.I)
    return int(m.group(1)) if m else None


def cable_slug(name: str) -> str | None:
    for prefix, slug_prefix in CABLE_PREFIXES:
        if prefix.lower() in name.lower():
            length = parse_length_m(name)
            if length:
                return f"{slug_prefix}-{length}m"
    return None


def parse_xlsx() -> tuple[dict[str, int], list[dict]]:
    ws = openpyxl.load_workbook(XLSX_PATH, data_only=True).active
    slug_prices: dict[str, int] = {}
    unmatched: list[dict] = []

    for r in range(12, ws.max_row + 1):
        raw_name = ws.cell(r, 1).value
        price = ws.cell(r, 3).value
        if not raw_name:
            continue
        name = norm_name(str(raw_name))
        if re.match(r"^\d+\.", name):
            continue
        if price in (None, "", "#REF!"):
            continue
        if isinstance(price, str) and price.strip().lower() == "договорная":
            unmatched.append({"name": name, "price": price, "reason": "договорная"})
            continue
        try:
            price_int = int(round(float(price)))
        except (TypeError, ValueError):
            continue

        lookup_name = name.rstrip(" .")
        slug = (
            SLUG_BY_EXCEL.get(name)
            or SLUG_BY_EXCEL.get(lookup_name)
            or cable_slug(name)
        )
        if slug:
            if slug not in slug_prices or price_int < slug_prices[slug]:
                slug_prices[slug] = price_int
        elif name.startswith("Комплект монтажных частей"):
            key = "montazhnoe-prisoedinitelnoe-oborudovanie"
            if key not in slug_prices or price_int < slug_prices[key]:
                slug_prices[key] = price_int
        else:
            unmatched.append({"name": name, "price": price_int, "desc": ws.cell(r, 2).value})

    return slug_prices, unmatched


def apply_seed(slug_prices: dict[str, int]) -> int:
    with open(SEED_PATH, encoding="utf-8") as f:
        content = f.read()

    updated = 0
    for slug, price in slug_prices.items():
        pattern = re.compile(
            rf'(slug:\s*"{re.escape(slug)}"[\s\S]*?price:\s*")([^"]+)(")',
            re.MULTILINE,
        )

        def repl(m: re.Match[str], p: str = str(price)) -> str:
            return m.group(1) + p + m.group(3)

        new_content, n = pattern.subn(repl, content, count=1)
        if n:
            content = new_content
            updated += 1

    with open(SEED_PATH, "w", encoding="utf-8", newline="\n") as f:
        f.write(content)
    return updated


def main() -> None:
    slug_prices, unmatched = parse_xlsx()
    updated = apply_seed(slug_prices)

    out = {
        "updated_slugs": updated,
        "slug_prices": slug_prices,
        "unmatched": unmatched,
    }
    out_path = os.path.join(os.path.dirname(__file__), "_technomer_applied.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)

    print(f"Updated {updated} products in seed.ts")
    print(f"Unmatched excel rows: {len(unmatched)}")
    for item in unmatched:
        print(f"  - {item['name']}: {item.get('price', item.get('reason'))}")


if __name__ == "__main__":
    main()
