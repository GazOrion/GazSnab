# -*- coding: utf-8 -*-
import json
import os
import re
import sys

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

ROOT = os.path.join(os.path.dirname(__file__), "..")
SEED_PATH = os.path.join(ROOT, "prisma", "seed.ts")
TARGET_PATH = os.path.join(ROOT, "public", "прайсы", "техномер.xlsx")


def normalize(text: object) -> str:
    if text is None:
        return ""
    s = str(text).strip().lower()
    s = s.replace("ё", "е")
    s = re.sub(r"\s+", " ", s)
    s = re.sub(r"[«»\"'„“]", "", s)
    return s


def parse_seed_products() -> list[dict]:
    with open(SEED_PATH, encoding="utf-8") as f:
        content = f.read()

    products = []
    # Match product blocks: title, slug, price
    pattern = re.compile(
        r'title:\s*"([^"]+)"[\s\S]*?slug:\s*"([^"]+)"[\s\S]*?price:\s*"([^"]+)"',
        re.MULTILINE,
    )
    for m in pattern.finditer(content):
        title, slug, price = m.groups()
        products.append({"title": title, "slug": slug, "price": price})

    return products


def parse_xlsx() -> list[dict]:
    wb = openpyxl.load_workbook(TARGET_PATH, data_only=True)
    items = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        name_col = 2
        price_col = 3
        header_row = 1
        for r in range(1, min(ws.max_row, 30) + 1):
            values = [normalize(ws.cell(r, c).value) for c in range(1, 8)]
            if any("наимен" in v for v in values):
                header_row = r
                for c in range(1, ws.max_column + 1):
                    cell = normalize(ws.cell(r, c).value)
                    if "наимен" in cell:
                        name_col = c
                break

        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, name_col).value
            price = ws.cell(r, price_col).value
            if not name or price in (None, ""):
                continue
            name_str = str(name).strip()
            if not name_str or name_str.lower().startswith("итого"):
                continue
            try:
                price_int = int(round(float(price)))
            except (TypeError, ValueError):
                continue
            items.append({"name": name_str, "price": price_int, "row": r})

    return items


def score_match(excel_name: str, product_title: str) -> float:
    a = normalize(excel_name)
    b = normalize(product_title)
    if a == b:
        return 100.0
    if a in b or b in a:
        return 90.0
    # token overlap
    ta = set(re.findall(r"[a-zа-я0-9]+", a))
    tb = set(re.findall(r"[a-zа-я0-9]+", b))
    if not ta or not tb:
        return 0.0
    inter = ta & tb
    union = ta | tb
    jaccard = len(inter) / len(union)
    # boost if model codes match
    codes_a = set(re.findall(r"[a-z]{1,3}[- ]?[a-z0-9]+", a))
    codes_b = set(re.findall(r"[a-z]{1,3}[- ]?[a-z0-9]+", b))
    if codes_a & codes_b:
        jaccard += 0.25
    return jaccard * 100


def main() -> None:
    products = parse_seed_products()
    items = parse_xlsx()

    matched = []
    unmatched_excel = []
    used_slugs = set()

    for item in items:
        best = None
        best_score = 0.0
        for p in products:
            s = score_match(item["name"], p["title"])
            if s > best_score:
                best_score = s
                best = p
        if best and best_score >= 55:
            matched.append(
                {
                    "excel_name": item["name"],
                    "product_title": best["title"],
                    "slug": best["slug"],
                    "old_price": best["price"],
                    "new_price": str(item["price"]),
                    "score": round(best_score, 1),
                }
            )
            used_slugs.add(best["slug"])
        else:
            unmatched_excel.append(item)

    out = {
        "matched_count": len(matched),
        "unmatched_count": len(unmatched_excel),
        "matched": matched,
        "unmatched_excel": unmatched_excel,
    }
    out_path = os.path.join(os.path.dirname(__file__), "_technomer_match.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(json.dumps(out, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
