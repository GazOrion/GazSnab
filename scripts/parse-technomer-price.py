import json
import os
import re
import sys

import openpyxl

PUBLIC = os.path.join(os.path.dirname(__file__), "..", "public")
PRICES_DIR = os.path.join(PUBLIC, "прайсы")
TARGET_NAME = "техномер.xlsx"


def normalize(text: object) -> str:
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text).strip().lower())


def main() -> None:
    target_path = os.path.join(PRICES_DIR, TARGET_NAME)
    if not os.path.exists(target_path):
        for root, _, files in os.walk(PUBLIC):
            for name in files:
                if name.lower() == TARGET_NAME:
                    target_path = os.path.join(root, name)
                    break

    if not os.path.exists(target_path):
        raise SystemExit(f"File not found: {TARGET_NAME}")

    wb = openpyxl.load_workbook(target_path, data_only=True)
    rows_out = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        header_row = None
        name_col = None
        price_col = 3  # column C

        for r in range(1, min(ws.max_row, 30) + 1):
            values = [ws.cell(r, c).value for c in range(1, min(ws.max_column, 12) + 1)]
            joined = " ".join(normalize(v) for v in values if v)
            if "наимен" in joined or "назван" in joined:
                header_row = r
                for c in range(1, ws.max_column + 1):
                    cell = normalize(ws.cell(r, c).value)
                    if "наимен" in cell or "назван" in cell or cell in {"товар", "продукция"}:
                        name_col = c
                break

        if header_row is None:
            name_col = 2 if ws.max_column >= 2 else 1
            header_row = 1

        for r in range(header_row + 1, ws.max_row + 1):
            name = ws.cell(r, name_col).value if name_col else ws.cell(r, 1).value
            price = ws.cell(r, price_col).value
            if not name:
                continue
            name_str = str(name).strip()
            if not name_str or name_str.lower().startswith("итого"):
                continue
            if price is None or price == "":
                continue
            if isinstance(price, str) and not re.search(r"\d", price):
                continue
            rows_out.append(
                {
                    "sheet": sheet_name,
                    "name": name_str,
                    "price": price,
                    "row": r,
                }
            )

    print(json.dumps({"file": target_path, "items": rows_out}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
